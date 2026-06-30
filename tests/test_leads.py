"""Тесты записи заявок в xlsx."""
from datetime import datetime

from openpyxl import load_workbook

from services.leads import HEADERS, Lead, append_lead


def _fixed_now():
    return datetime(2026, 6, 29, 14, 30)


def test_creates_file_with_headers(tmp_path):
    path = tmp_path / "leads.xlsx"
    append_lead(Lead("Иван", "+79991234567", "суббота"), path=path, now=_fixed_now)

    assert path.exists()
    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == HEADERS
    assert [c.value for c in ws[2]] == ["2026-06-29 14:30", "Иван", "+79991234567", "суббота"]


def test_appends_without_overwriting(tmp_path):
    path = tmp_path / "leads.xlsx"
    append_lead(Lead("Иван", "+700"), path=path, now=_fixed_now)
    append_lead(Lead("Пётр", "@petr"), path=path, now=_fixed_now)

    ws = load_workbook(path).active
    assert ws.max_row == 3  # заголовок + 2 заявки
    assert ws[3][1].value == "Пётр"


def test_default_time_is_dash(tmp_path):
    path = tmp_path / "leads.xlsx"
    append_lead(Lead("Аня", "@anya"), path=path, now=_fixed_now)

    ws = load_workbook(path).active
    assert ws[2][3].value == "—"
