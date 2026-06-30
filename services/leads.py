"""Хранение заявок: дозапись в data/leads.xlsx (openpyxl).

Файл создаётся при первой заявке с заголовками. Каждая заявка — новая строка.
Дубль заявки уходит владельцу в Telegram (см. services/notify.py) — здесь только
надёжное локальное хранилище, которое не зависит от доставки сообщений.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook, load_workbook

from core.config import LEADS_XLSX

HEADERS = ["Дата", "Имя", "Контакт", "Удобное время"]
SHEET_TITLE = "Заявки"


@dataclass(frozen=True)
class Lead:
    """Одна собранная заявка."""
    name: str
    contact: str
    preferred_time: str = "—"

    def as_row(self, created_at: str) -> list[str]:
        return [created_at, self.name, self.contact, self.preferred_time]


def append_lead(
    lead: Lead,
    path: Path = LEADS_XLSX,
    now: Optional[Callable[[], datetime]] = None,
) -> Path:
    """Дозаписывает заявку в xlsx. Создаёт файл с заголовками, если его нет.

    `now` — фабрика времени (для тестов). Возвращает путь к файлу.
    """
    now = now or datetime.now
    created_at = now().strftime("%Y-%m-%d %H:%M")

    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_TITLE
        ws.append(HEADERS)

    ws.append(lead.as_row(created_at))
    wb.save(path)
    return path
